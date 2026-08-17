"""Contrôles des livrables : le classeur, les fichiers tabulaires, la feuille de dictionnaire.

**Exigence de méthode tenue partout ici : chaque contrôle lit le fichier produit.** Interroger la
base pour vérifier un export ne vérifierait rien de l'export — c'est le défaut que plusieurs
mutations ont déjà révélé ailleurs, sous d'autres formes : un contrôle qui n'observe pas ce qu'il
prétend observer. La base n'intervient donc que comme second membre d'une égalité, jamais comme
substitut du livrable.

Aucun travail au niveau du module : ni connexion, ni ouverture de classeur à l'import. Le fichier
se collecte sur un clone frais sans base ni variable exportée.

Aucun littéral de volumétrie. Chaque attendu est une égalité entre deux mesures — ce que porte le
fichier d'un côté, ce que porte l'instantané de l'autre.
"""

from __future__ import annotations

import csv
import subprocess
import warnings
from pathlib import Path

import pytest

RACINE = Path(__file__).resolve().parent.parent
SORTIE = RACINE / "exports"


def _exporter():
    """Importé à l'appel, jamais à l'import : le module ouvre des connexions."""
    from livraison import exporter

    return exporter


def _normaliser(valeur) -> str:
    """Rend une valeur sous une forme comparable de part et d'autre du format de classeur.

    Le format ne conserve pas tous les types du serveur, et la mesure l'a établi en confrontant
    une ligne cellule par cellule : une **date** en ressort en horodatage à minuit, une **chaîne
    vide** en ressort **nulle**, un **décimal exact** en ressort en nombre à virgule flottante dont
    le dernier chiffre significatif peut différer. Ce sont des propriétés de l'aller-retour, non des
    altérations de l'export.

    La normalisation s'applique **identiquement aux deux membres** de la comparaison : elle efface
    ces différences de représentation et **aucune autre**, si bien qu'une valeur réellement
    changée reste distinguée — ce qu'une mutation doit prouver.
    """
    from datetime import date, datetime
    from decimal import Decimal

    if valeur is None or valeur == "":
        return "∅"
    if isinstance(valeur, datetime):
        return valeur.replace(tzinfo=None).isoformat()
    if isinstance(valeur, date):
        return datetime(valeur.year, valeur.month, valeur.day).isoformat()
    if isinstance(valeur, Decimal | float):
        # Le format sérialise les nombres en double précision et n'en restitue que quinze
        # chiffres significatifs : mesuré sur les 24 692 cellules numériques du classeur, l'écart
        # relatif maximal de l'aller-retour vaut 5,24 × 10⁻¹⁶. Douze chiffres significatifs sont
        # donc trois ordres de grandeur au-dessus de ce bruit, et très en deçà de ce que
        # produirait une valeur réellement altérée.
        return f"{float(valeur):.12g}"
    return str(valeur)


def _empreinte(valeurs) -> int:
    """Empreinte agrégée d'un ensemble de lignes, indépendante de leur ordre.

    Chaque ligne est condensée, puis les condensés sont sommés : l'addition étant commutative,
    l'empreinte ne dépend pas de l'ordre des lignes. C'est la même famille d'empreinte que celle
    employée pour l'égalité de l'instantané, transposée hors du serveur puisqu'il faut ici comparer
    un fichier à une table.
    """
    import hashlib

    total = 0
    for ligne in valeurs:
        rendu = "\x1f".join(_normaliser(v) for v in ligne)
        total += int(hashlib.md5(rendu.encode("utf-8")).hexdigest()[:8], 16)
    return total


@pytest.fixture(scope="session")
def livrables():
    """Précondition établie par les contrôles : l'instantané est constitué, puis l'export produit.

    Rien n'est hérité d'une exécution antérieure — ni l'instantané, ni les fichiers.
    """
    exporter = _exporter()
    try:
        conn = exporter._connexion()
        conn.close()
    except Exception as echec:  # pragma: no cover - dépend de l'environnement
        pytest.fail(f"la base n'est pas joignable : {echec}")

    from instantane import rafraichir

    conn = exporter._connexion()
    try:
        with conn.cursor() as curseur:
            situation = exporter.etat(curseur)
    finally:
        conn.close()
    if situation["rafraichi_le"] is None:
        reussite, message = rafraichir.rafraichir()
        if not reussite:
            pytest.fail(f"l'instantané ne peut être constitué : {message}")

    reussite, message = exporter.exporter()
    if not reussite:
        pytest.fail(f"l'export a échoué, les contrôles n'ont rien à lire : {message}")
    return SORTIE / exporter.NOM_CLASSEUR


def _tables_du_catalogue(prefixes: tuple[str, ...] | None = None) -> list[str]:
    """La liste des tables, interrogée par le contrôle lui-même.

    Elle n'est PAS obtenue par les fonctions du module exporté : les deux membres de l'égalité
    viendraient alors de la même source, et une omission dans le module se répercuterait dans
    l'attendu. C'est le défaut qu'une mutation a mis au jour ici même.
    """
    exporter = _exporter()
    conn = exporter._connexion()
    try:
        with conn.cursor() as curseur:
            curseur.execute(
                "select c.relname from pg_class c "
                "join pg_namespace n on n.oid = c.relnamespace "
                "where n.nspname = %s and c.relkind = 'r' order by c.relname",
                (exporter.SCHEMA,),
            )
            tables = [ligne[0] for ligne in curseur.fetchall()]
    finally:
        conn.close()
    if prefixes is None:
        return tables
    return [table for table in tables if table.startswith(prefixes)]


def _table_de_l_instantane(nom: str):
    exporter = _exporter()
    conn = exporter._connexion()
    try:
        with conn.cursor() as curseur:
            colonnes = [c for c, _ in exporter._colonnes(curseur, nom)]
            curseur.execute(f"select * from {nom}")
            return colonnes, curseur.fetchall()
    finally:
        conn.close()


def test_le_classeur_s_ouvre_sans_avertissement(livrables) -> None:
    """Les avertissements sont capturés, non supposés absents.

    Se fier au silence apparent ne prouve rien : un avertissement peut être émis sans être affiché
    selon la configuration. Ils sont donc interceptés et comptés.
    """
    import openpyxl

    with warnings.catch_warnings(record=True) as captures:
        warnings.simplefilter("always")
        classeur = openpyxl.load_workbook(livrables, read_only=True)
        classeur.close()

    emis = [f"{c.category.__name__}: {c.message}" for c in captures]
    assert not emis, f"avertissements à l'ouverture du classeur : {emis}"


def test_le_classeur_porte_exactement_les_feuilles_attendues(livrables) -> None:
    """Le nombre de feuilles est calculé des deux côtés, jamais écrit."""
    import openpyxl

    exporter = _exporter()
    classeur = openpyxl.load_workbook(livrables, read_only=True)
    try:
        feuilles = list(classeur.sheetnames)
    finally:
        classeur.close()

    attendues = _tables_du_catalogue(("dim_", "fct_"))

    assert len(feuilles) == len(attendues) + 1, (
        f"{len(feuilles)} feuilles au classeur pour {len(attendues)} tables du schéma en étoile "
        "plus le dictionnaire"
    )
    manquantes = sorted(set(attendues) - set(feuilles))
    assert not manquantes, f"tables du schéma en étoile absentes du classeur : {manquantes}"
    en_trop = sorted(set(feuilles) - set(attendues) - {exporter.FEUILLE_DICTIONNAIRE})
    assert not en_trop, f"feuilles du classeur sans table correspondante : {en_trop}"


def test_chaque_feuille_a_les_dimensions_de_sa_table(livrables) -> None:
    """Lignes et colonnes de la feuille contre celles de la table, table par table."""
    import openpyxl

    exporter = _exporter()
    classeur = openpyxl.load_workbook(livrables, read_only=True)
    ecarts = []
    try:
        for feuille in classeur.worksheets:
            if feuille.title == exporter.FEUILLE_DICTIONNAIRE:
                continue
            colonnes, lignes = _table_de_l_instantane(feuille.title)
            if feuille.max_row != len(lignes) + 1:
                ecarts.append(
                    f"{feuille.title} : {feuille.max_row} lignes au classeur contre "
                    f"{len(lignes) + 1} attendues (en-tête comprise)"
                )
            if feuille.max_column != len(colonnes):
                ecarts.append(
                    f"{feuille.title} : {feuille.max_column} colonnes contre {len(colonnes)}"
                )
    finally:
        classeur.close()

    assert not ecarts, "dimensions divergentes : " + " | ".join(ecarts)


def test_chaque_feuille_a_le_contenu_de_sa_table(livrables) -> None:
    """Au-delà des dimensions : une empreinte du contenu, indépendante de l'ordre.

    Deux feuilles de mêmes dimensions peuvent porter des valeurs différentes ; c'est ce que cette
    propriété distingue, et c'est à une mutation de le prouver.
    """
    import openpyxl

    exporter = _exporter()
    classeur = openpyxl.load_workbook(livrables, read_only=True)
    ecarts = []
    try:
        for feuille in classeur.worksheets:
            if feuille.title == exporter.FEUILLE_DICTIONNAIRE:
                continue
            colonnes, lignes = _table_de_l_instantane(feuille.title)
            rangs = list(feuille.iter_rows(values_only=True))
            entete, corps = rangs[0], rangs[1:]

            if list(entete) != colonnes:
                ecarts.append(f"{feuille.title} : en-tête différent des colonnes de la table")
                continue

            attendu = _empreinte([[exporter._pour_classeur(v) for v in ligne] for ligne in lignes])
            obtenu = _empreinte(corps)
            if obtenu != attendu:
                ecarts.append(f"{feuille.title} : empreinte {obtenu} contre {attendu}")
    finally:
        classeur.close()

    assert not ecarts, "contenus divergents à dimensions possiblement égales : " + " | ".join(
        ecarts
    )


def test_les_fichiers_tabulaires_couvrent_les_tables_de_l_instantane(livrables) -> None:
    """Couverture dans les deux sens, et nombre de lignes de chacun."""
    tables = _tables_du_catalogue()

    presents = {chemin.stem for chemin in SORTIE.glob("*.csv")}
    manquants = sorted(set(tables) - presents)
    assert not manquants, f"tables sans fichier tabulaire : {manquants}"
    en_trop = sorted(presents - set(tables))
    assert not en_trop, f"fichiers tabulaires sans table correspondante : {en_trop}"

    ecarts = []
    for table in tables:
        _, lignes = _table_de_l_instantane(table)
        with (SORTIE / f"{table}.csv").open(encoding="utf-8-sig", newline="") as fichier:
            comptees = sum(1 for _ in csv.reader(fichier)) - 1
        if comptees != len(lignes):
            ecarts.append(f"{table} : {comptees} lignes au fichier contre {len(lignes)}")
    assert not ecarts, "décomptes divergents : " + " | ".join(ecarts)


def test_le_dictionnaire_couvre_les_colonnes_exportees(livrables) -> None:
    """Couverture dans les deux sens, et somme des documentées et non documentées."""
    import openpyxl

    exporter = _exporter()
    classeur = openpyxl.load_workbook(livrables, read_only=True)
    try:
        rangs = list(classeur[exporter.FEUILLE_DICTIONNAIRE].iter_rows(values_only=True))
    finally:
        classeur.close()

    corps = [ligne for ligne in rangs[3:] if ligne and ligne[0]]
    couples_ecrits = {(ligne[0], ligne[1]) for ligne in corps}

    conn = exporter._connexion()
    try:
        with conn.cursor() as curseur:
            attendus = {
                (table, colonne)
                for table in _tables_du_catalogue(("dim_", "fct_"))
                for colonne, _ in exporter._colonnes(curseur, table)
            }
    finally:
        conn.close()

    manquants = sorted(attendus - couples_ecrits)
    assert not manquants, f"colonnes exportées absentes du dictionnaire : {manquants[:5]}"
    en_trop = sorted(couples_ecrits - attendus)
    assert not en_trop, f"lignes du dictionnaire sans colonne exportée : {en_trop[:5]}"

    sans = sum(1 for ligne in corps if ligne[3] == exporter.MENTION_SANS_DESCRIPTION)
    avec = len(corps) - sans
    assert avec + sans == len(attendus), (
        f"{avec} documentées et {sans} non documentées font {avec + sans} pour "
        f"{len(attendus)} colonnes exportées"
    )

    # Ce que porte chaque ligne est confronté à ce que les fichiers d'accompagnement DÉCLARENT :
    # une colonne sans description déclarée doit porter la mention d'absence, et non un texte
    # inventé qui la ferait passer pour documentée. Sans cette confrontation, combler les absences
    # par une phrase quelconque laisserait le décompte inchangé et le contrôle vert.
    declare = exporter.descriptions_declarees()
    fautifs = []
    for table, colonne, _type, description, _table_desc in corps:
        attendue = declare.get(table, {}).get("colonnes", {}).get(colonne, "")
        if attendue and description != attendue:
            fautifs.append(f"{table}.{colonne} : description différente de celle déclarée")
        if not attendue and description != exporter.MENTION_SANS_DESCRIPTION:
            fautifs.append(f"{table}.{colonne} : « {description[:40]} » là où rien n'est déclaré")
    assert not fautifs, "descriptions non conformes aux déclarations : " + " | ".join(fautifs[:5])


def test_la_date_de_reference_du_dictionnaire_est_celle_des_donnees(livrables) -> None:
    """La date écrite dans le classeur, contre la table d'état, contre la couche source."""
    import openpyxl

    exporter = _exporter()
    classeur = openpyxl.load_workbook(livrables, read_only=True)
    try:
        entete = list(classeur[exporter.FEUILLE_DICTIONNAIRE].iter_rows(values_only=True))[0]
    finally:
        classeur.close()

    ecrite = entete[1]

    conn = exporter._connexion()
    try:
        with conn.cursor() as curseur:
            portee = exporter.etat(curseur)["date_reference"]
            curseur.execute(
                "select table_name from information_schema.columns "
                "where table_schema = 'source' and column_name = 'date_extraction'"
            )
            tables = [ligne[0] for ligne in curseur.fetchall()]
            union = " union all ".join(
                f"select max(to_date(date_extraction, 'MM/DD/YYYY')) as d from source.{t}"
                for t in tables
            )
            curseur.execute(f"select max(d) from ({union}) as toutes")
            mesuree = curseur.fetchone()[0]
    finally:
        conn.close()

    assert ecrite == f"{portee:%d/%m/%Y}", (
        f"le dictionnaire porte « {ecrite} » là où la table d'état porte {portee}"
    )
    assert portee == mesuree, (
        f"la table d'état porte {portee}, la dernière extraction chargée est {mesuree}"
    )


def test_l_export_ne_lit_que_l_instantane(livrables) -> None:
    """Garantie structurelle, éprouvée : une lecture non qualifiée d'une autre couche échoue."""
    exporter = _exporter()
    conn = exporter._connexion()
    try:
        with conn.cursor() as curseur:
            curseur.execute("select count(*) from fct_sejour")
            assert curseur.fetchone()[0] > 0, "l'instantané paraît vide"

            # Le chemin de recherche est lu tel que la connexion le porte : c'est lui qui décide
            # ce qu'un nom non qualifié désigne. La couche `marts` porte les MÊMES noms de table
            # que l'instantané ; l'y ajouter ferait lire des vues reconstruites chaque jour à la
            # place de l'état figé, sans qu'aucune requête n'échoue. Une propriété qui se
            # contenterait de vérifier qu'une couche aux noms différents est hors d'atteinte ne
            # verrait rien de cette substitution.
            curseur.execute("show search_path")
            chemin = curseur.fetchone()[0]

            with pytest.raises(Exception) as capture:
                curseur.execute("select count(*) from patients")
    finally:
        conn.close()

    couches = {partie.strip().strip('"') for partie in chemin.split(",")}
    assert couches == {exporter.SCHEMA}, (
        f"le chemin de recherche porte {sorted(couches)} : une couche autre que l'instantané "
        "peut être lue sans qualification"
    )
    assert "does not exist" in str(capture.value), (
        f"une table d'une autre couche a été atteinte sans qualification : {capture.value}"
    )


def test_rien_du_repertoire_de_sortie_n_est_suivi(livrables) -> None:
    """Le contrôle est éprouvé contre un cas positif avant que son silence ne soit cru."""
    suivis = subprocess.run(
        ["git", "ls-files", "exports/"],
        capture_output=True,
        text=True,
        cwd=RACINE,
        check=True,
    ).stdout.split()

    # Cas positif : le fichier de marque EST suivi. Si la commande ne le trouvait pas, son silence
    # sur les autres fichiers ne prouverait rien.
    assert "exports/.gitkeep" in suivis, (
        "le fichier de marque n'est pas trouvé comme suivi : la recherche ne prouve rien"
    )

    produits = sorted(set(suivis) - {"exports/.gitkeep"})
    assert not produits, f"des fichiers d'export sont suivis : {produits}"

    # Et rien n'est écrit HORS du répertoire de sortie : un export qui déposerait ses fichiers
    # ailleurs — à la racine, par exemple — ne serait vu ni par la règle d'exclusion ni par la
    # recherche ci-dessus, et salirait le dépôt sans que rien ne le signale.
    exporter = _exporter()
    conn = exporter._connexion()
    try:
        with conn.cursor() as curseur:
            tables = exporter.tables_de_l_instantane(curseur)
    finally:
        conn.close()

    egares = [chemin.name for chemin in RACINE.glob("*.csv") if chemin.stem in tables]
    assert not egares, (
        f"des fichiers d'export ont été écrits hors du répertoire de sortie : {egares}"
    )


def test_deux_exports_successifs_produisent_le_meme_contenu(livrables) -> None:
    """Comparaison des CONTENUS, jamais des octets.

    Un classeur porte des métadonnées d'horodatage : deux fichiers produits à une seconde
    d'intervalle ne sont pas identiques octet à octet, et une propriété d'identité binaire
    rougirait sans cause. Ce sont les valeurs des feuilles qui sont comparées.
    """
    import openpyxl

    exporter = _exporter()

    def empreintes_du_classeur(chemin: Path) -> dict[str, int]:
        classeur = openpyxl.load_workbook(chemin, read_only=True)
        try:
            return {
                feuille.title: _empreinte(feuille.iter_rows(values_only=True))
                for feuille in classeur.worksheets
            }
        finally:
            classeur.close()

    avant = empreintes_du_classeur(livrables)

    reussite, message = exporter.exporter()
    assert reussite, message

    apres = empreintes_du_classeur(livrables)

    assert set(avant) == set(apres), (
        f"les feuilles diffèrent entre deux exports : {set(avant) ^ set(apres)}"
    )
    divergentes = sorted(nom for nom in avant if avant[nom] != apres[nom])
    # La feuille de dictionnaire porte l'horodatage du rafraîchissement : elle ne diffère que si
    # celui-ci a changé entre les deux exports, ce que ce contrôle ne provoque pas.
    assert not divergentes, f"feuilles au contenu différent entre deux exports : {divergentes}"
