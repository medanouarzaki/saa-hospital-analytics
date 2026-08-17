"""Produit les livrables : un classeur à lire, des fichiers tabulaires à réutiliser.

Le répertoire de ce module s'appelle `livraison` et non `export` : le répertoire de SORTIE
s'appelle `exports`, et deux noms voisins auraient fini par être confondus dans une commande ou un
chemin. Le module produit, le répertoire reçoit.

**Tout est lu dans le schéma d'instantané, jamais dans la couche `marts`.** C'est ce qui garantit
qu'un export et un écran ouverts au même moment portent le même état : les vues de la chaîne se
reconstruisent chaque jour, l'instantané non. La restriction est structurelle, comme pour le
tableau de bord — chaque connexion réduit son chemin de recherche au seul schéma d'instantané, si
bien qu'une requête nommant sans le qualifier un objet d'une autre couche échoue à l'exécution.

Deux formes, deux destinataires. **Le classeur** porte une feuille par table du schéma en étoile —
six dimensions, six faits — plus une feuille de dictionnaire ; il se lit. **Les fichiers
tabulaires** portent toutes les tables de l'instantané, une par fichier ; ils se réutilisent, et
rien de ce que la chaîne produit n'est perdu à l'export.

Les fichiers tabulaires sont écrits en encodage universel **avec marque d'ordre d'octets** et
séparateur virgule. La marque est là pour qu'un tableur ouvrant le fichier par double-clic
reconnaisse l'encodage et n'abîme pas les caractères accentués, qui abondent dans les libellés.
Limite connue et assumée : un tableur configuré avec la virgule comme séparateur décimal — usage
courant en France — lira mal un fichier à séparateur virgule et demandera un import explicite. Le
classeur existe précisément pour le lecteur qui ne veut pas s'en soucier.
"""

from __future__ import annotations

import argparse
import csv
import sys
import time
from datetime import UTC, datetime
from pathlib import Path

import psycopg
import yaml
from openpyxl import Workbook

# Limites d'une feuille, lues dans la bibliothèque installée plutôt que recopiées, de sorte qu'une
# montée de version qui les changerait soit suivie sans intervention.
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW

from ingestion import appliquer_ddl

RACINE = Path(__file__).resolve().parent.parent
SORTIE = RACINE / "exports"
MODELES_MARTS = RACINE / "dbt" / "models" / "marts"

SCHEMA = "instantane"
TABLE_ETAT = "instantane_etat"

NOM_CLASSEUR = "tableau_de_bord.xlsx"
FEUILLE_DICTIONNAIRE = "Dictionnaire"

# Les tables du schéma en étoile, seules à figurer dans le classeur. Elles se reconnaissent à leur
# préfixe, que la chaîne applique sans exception ; la liste n'est donc pas écrite à la main.
PREFIXES_ETOILE = ("dim_", "fct_")

MENTION_SANS_DESCRIPTION = "Non documentée"


class ExportImpossible(RuntimeError):
    """L'export ne peut être produit ; le message dit pourquoi."""


def _connexion():
    """Connexion dont le chemin de recherche est réduit au seul schéma d'instantané.

    Même configuration que le reste du dépôt : les variables sont lues par le chargeur
    d'environnement partagé, et aucun second chemin n'est inventé.
    """
    variables = appliquer_ddl.charger_environnement()
    conn = psycopg.connect(
        host=variables["POSTGRES_HOST"],
        port=variables["POSTGRES_PORT"],
        dbname=variables["POSTGRES_DB"],
        user=variables["POSTGRES_USER"],
        password=variables.get("POSTGRES_PASSWORD", ""),
    )
    with conn.cursor() as curseur:
        curseur.execute(f"set search_path to {SCHEMA}")
        curseur.execute("set time zone 'UTC'")
    return conn


def tables_de_l_instantane(curseur) -> list[str]:
    curseur.execute(
        "select c.relname from pg_class c join pg_namespace n on n.oid = c.relnamespace "
        f"where n.nspname = '{SCHEMA}' and c.relkind = 'r' order by c.relname"
    )
    return [ligne[0] for ligne in curseur.fetchall()]


def tables_du_schema_en_etoile(curseur) -> list[str]:
    return [table for table in tables_de_l_instantane(curseur) if table.startswith(PREFIXES_ETOILE)]


def etat(curseur) -> dict:
    curseur.execute(f"select max(rafraichi_le), max(date_reference_donnees) from {TABLE_ETAT}")
    rafraichi_le, date_reference = curseur.fetchone()
    return {"rafraichi_le": rafraichi_le, "date_reference": date_reference}


def _colonnes(curseur, table: str) -> list[tuple[str, str]]:
    curseur.execute(
        "select column_name, data_type from information_schema.columns "
        "where table_schema = %s and table_name = %s order by ordinal_position",
        (SCHEMA, table),
    )
    return list(curseur.fetchall())


def descriptions_declarees() -> dict[str, dict]:
    """Ce que les fichiers d'accompagnement des modèles déclarent, table par table.

    Rien n'est inventé : ce qui n'est pas déclaré ressort comme non déclaré, et la feuille de
    dictionnaire le dit à la place de le combler.
    """
    declare: dict[str, dict] = {}
    for fichier in sorted(MODELES_MARTS.glob("*.yml")):
        contenu = yaml.safe_load(fichier.read_text(encoding="utf-8")) or {}
        for modele in contenu.get("models", []):
            declare[modele["name"]] = {
                "table": " ".join((modele.get("description") or "").split()),
                "colonnes": {
                    colonne["name"]: " ".join((colonne.get("description") or "").split())
                    for colonne in modele.get("columns", [])
                },
            }
    return declare


def lignes_du_dictionnaire(curseur, tables: list[str]) -> list[list[str]]:
    """Une ligne par couple table-colonne des tables EXPORTÉES, et d'aucune autre population."""
    declare = descriptions_declarees()
    lignes = []
    for table in tables:
        declaration = declare.get(table, {})
        description_table = declaration.get("table", "") or MENTION_SANS_DESCRIPTION
        for colonne, type_sql in _colonnes(curseur, table):
            description = declaration.get("colonnes", {}).get(colonne, "")
            lignes.append(
                [
                    table,
                    colonne,
                    type_sql,
                    description or MENTION_SANS_DESCRIPTION,
                    description_table,
                ]
            )
    return lignes


def _pour_classeur(valeur):
    """Adapte une valeur au format du classeur, qui n'accepte pas tous les types du serveur.

    Le format de classeur **refuse un horodatage portant un fuseau** — la bibliothèque lève plutôt
    que d'écrire. La connexion étant fixée à UTC, retirer le fuseau ne déplace aucun instant : la
    valeur écrite est l'instant UTC, et l'en-tête du dictionnaire le dit au lecteur.
    """
    if isinstance(valeur, datetime) and valeur.tzinfo is not None:
        return valeur.astimezone(UTC).replace(tzinfo=None)
    return "" if valeur is None else valeur


def _ecrire_feuille(feuille, curseur, table: str) -> int:
    """Écrit l'en-tête puis les lignes ; rend le nombre de cellules écrites."""
    colonnes = [nom for nom, _ in _colonnes(curseur, table)]
    feuille.append(colonnes)
    curseur.execute(f"select * from {table}")
    lignes = curseur.fetchall()
    for ligne in lignes:
        feuille.append([_pour_classeur(valeur) for valeur in ligne])
    feuille.freeze_panes = "A2"
    return (len(lignes) + 1) * len(colonnes)


def produire_classeur(destination: Path | None = None) -> dict:
    """Le classeur : une feuille par table du schéma en étoile, plus le dictionnaire."""
    depart = time.monotonic()
    destination = destination or (SORTIE / NOM_CLASSEUR)
    destination.parent.mkdir(parents=True, exist_ok=True)

    classeur = Workbook()
    classeur.remove(classeur.active)
    cellules = 0

    conn = _connexion()
    try:
        with conn.cursor() as curseur:
            tables = tables_du_schema_en_etoile(curseur)
            if not tables:
                raise ExportImpossible("aucune table du schéma en étoile dans l'instantané")

            situation = etat(curseur)
            for table in tables:
                cellules += _ecrire_feuille(classeur.create_sheet(table), curseur, table)

            feuille = classeur.create_sheet(FEUILLE_DICTIONNAIRE)
            # L'en-tête porte la date des données et celle du rafraîchissement : un classeur
            # circule sans son contexte, et sans cette mention un lecteur le croira du jour.
            feuille.append(
                [
                    "Données arrêtées au",
                    f"{situation['date_reference']:%d/%m/%Y}",
                    "État constitué le",
                    f"{situation['rafraichi_le']:%d/%m/%Y à %H:%M} (UTC)",
                ]
            )
            feuille.append([])
            feuille.append(
                ["Table", "Colonne", "Type", "Description de la colonne", "Description de la table"]
            )
            lignes = lignes_du_dictionnaire(curseur, tables)
            for ligne in lignes:
                feuille.append(ligne)
            feuille.freeze_panes = "A4"
            cellules += (len(lignes) + 3) * 5
    finally:
        conn.close()

    depassements = [
        feuille.title
        for feuille in classeur.worksheets
        if feuille.max_row > MAX_ROW or feuille.max_column > MAX_COLUMN
    ]
    if depassements:
        raise ExportImpossible(f"feuilles dépassant les limites du format : {depassements}")

    classeur.save(destination)
    return {
        "chemin": destination,
        "feuilles": len(classeur.worksheets),
        "cellules": cellules,
        "octets": destination.stat().st_size,
        "duree_s": time.monotonic() - depart,
    }


def produire_fichiers_tabulaires(destination: Path | None = None) -> dict:
    """Un fichier par table de l'instantané : rien de ce que la chaîne produit n'est perdu."""
    depart = time.monotonic()
    destination = destination or SORTIE
    destination.mkdir(parents=True, exist_ok=True)

    ecrits = []
    conn = _connexion()
    try:
        with conn.cursor() as curseur:
            for table in tables_de_l_instantane(curseur):
                colonnes = [nom for nom, _ in _colonnes(curseur, table)]
                curseur.execute(f"select * from {table}")
                lignes = curseur.fetchall()
                chemin = destination / f"{table}.csv"
                # Encodage universel AVEC marque d'ordre d'octets : un tableur ouvrant le fichier
                # par double-clic reconnaît alors l'encodage et n'abîme pas les accents.
                with chemin.open("w", encoding="utf-8-sig", newline="") as fichier:
                    plume = csv.writer(fichier, delimiter=",")
                    plume.writerow(colonnes)
                    for ligne in lignes:
                        plume.writerow(["" if v is None else v for v in ligne])
                ecrits.append((table, len(lignes), chemin.stat().st_size))
    finally:
        conn.close()

    # Les fichiers d'une production antérieure qui ne correspondent plus à aucune table sont
    # retirés. Sans cela, une table disparue de l'instantané laisserait indéfiniment son fichier,
    # qu'un destinataire prendrait pour un livrable à jour — et un contrôle de couverture le
    # trouverait présent alors qu'il n'a pas été écrit.
    ecrits_maintenant = {table for table, _, _ in ecrits}
    perimes = [
        chemin for chemin in destination.glob("*.csv") if chemin.stem not in ecrits_maintenant
    ]
    for chemin in perimes:
        chemin.unlink()

    return {
        "fichiers": len(ecrits),
        "perimes_retires": len(perimes),
        "lignes": sum(n for _, n, _ in ecrits),
        "octets": sum(o for _, _, o in ecrits),
        "duree_s": time.monotonic() - depart,
        "detail": ecrits,
    }


def exporter() -> tuple[bool, str]:
    """Renvoie (réussite, message), à la manière des autres étapes de la chaîne."""
    try:
        classeur = produire_classeur()
        tabulaires = produire_fichiers_tabulaires()
    except (ExportImpossible, psycopg.Error) as echec:
        return False, f"export : ECHEC - {echec}"

    return True, (
        f"export : OK - classeur {classeur['feuilles']} feuilles, {classeur['cellules']} cellules, "
        f"{classeur['octets']} octets en {classeur['duree_s']:.2f}s ; "
        f"{tabulaires['fichiers']} fichiers tabulaires, {tabulaires['lignes']} lignes, "
        f"{tabulaires['octets']} octets en {tabulaires['duree_s']:.2f}s"
    )


def main() -> None:
    analyseur = argparse.ArgumentParser(
        description="Produit les livrables du tableau de bord : un classeur des tables du schema "
        "en etoile avec son dictionnaire, et un fichier tabulaire par table de l'instantane."
    )
    analyseur.parse_args()

    reussite, message = exporter()
    print(message)
    if not reussite:
        sys.exit(1)


if __name__ == "__main__":
    main()
